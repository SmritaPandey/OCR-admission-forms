import logging

from fastapi import APIRouter, HTTPException, Depends, Query, Body
from sqlalchemy.orm import Session
from sqlalchemy import or_, and_, desc, asc
from typing import Optional, List, Dict, Any
from datetime import datetime, timedelta
import csv
import json
import io
from backend.database import get_db, AdmissionForm, FormStatus, StudentDocument
from backend.models.form import (
    FormResponse,
    FormDetailResponse,
    FormVerification,
    FormSearchParams,
    FormExtractionResponse,
    ExtractedData,
)
from backend.api.routes.students import get_or_create_student_profile
from backend.ocr import get_ocr_provider
from backend.utils.file_handler import load_image
from backend.config import settings
from backend.api.dependencies import RequireAnyAuth, RequireStaffOrAdmin
from backend.models.auth_models import CurrentUser

router = APIRouter()
logger = logging.getLogger(__name__)


def apply_form_filters(
    query,
    *,
    student_name: Optional[str] = None,
    phone_number: Optional[str] = None,
    email: Optional[str] = None,
    enrollment_number: Optional[str] = None,
    application_number: Optional[str] = None,
    course_applied: Optional[str] = None,
    status: Optional[FormStatus] = None,
    date_from: Optional[datetime] = None,
    date_to: Optional[datetime] = None,
) :
    """Apply common filters to the admission forms query."""
    if student_name:
        query = query.filter(AdmissionForm.student_name.ilike(f"%{student_name.strip()}%"))
    if phone_number:
        query = query.filter(AdmissionForm.phone_number.ilike(f"%{phone_number.strip()}%"))
    if email:
        query = query.filter(AdmissionForm.email.ilike(f"%{email.strip()}%"))
    if enrollment_number:
        query = query.filter(AdmissionForm.enrollment_number.ilike(f"%{enrollment_number.strip()}%"))
    if application_number:
        query = query.filter(AdmissionForm.application_number.ilike(f"%{application_number.strip()}%"))
    if course_applied:
        query = query.filter(AdmissionForm.course_applied.ilike(f"%{course_applied.strip()}%"))
    if status:
        query = query.filter(AdmissionForm.status == status)

    if date_from:
        start = date_from.replace(hour=0, minute=0, second=0, microsecond=0)
        query = query.filter(AdmissionForm.upload_date >= start)
    if date_to:
        end = date_to.replace(hour=23, minute=59, second=59, microsecond=999999)
        query = query.filter(AdmissionForm.upload_date <= end)

    return query

# --- Export Related Logic ---

EXPORT_FIELDS = [
    ("id", "Form ID"),
    ("filename", "Filename"),
    ("status", "Status"),
    ("student_profile_id", "Student Profile ID"),
    ("upload_date", "Upload Date"),
    ("verified_date", "Verified Date"),
    ("verified_by", "Verified By"),
    # Basic Details
    ("student_name", "Student Name"),
    ("date_of_birth", "Date of Birth"),
    ("gender", "Gender"),
    ("category", "Category"),
    ("nationality", "Nationality"),
    ("religion", "Religion"),
    ("aadhar_number", "Aadhar Number"),
    ("blood_group", "Blood Group"),
    # Address Details
    ("permanent_address", "Permanent Address"),
    ("correspondence_address", "Correspondence Address"),
    ("city", "City"),
    ("state", "State"),
    ("pincode", "Pincode"),
    # Contact Details
    ("phone_number", "Phone Number"),
    ("alternate_phone", "Alternate Phone"),
    ("email", "Email"),
    ("emergency_contact_name", "Emergency Contact Name"),
    ("emergency_contact_phone", "Emergency Contact Phone"),
    # Guardian / Parent Details
    ("father_name", "Father Name"),
    ("father_occupation", "Father Occupation"),
    ("father_phone", "Father Phone"),
    ("mother_name", "Mother Name"),
    ("mother_occupation", "Mother Occupation"),
    ("mother_phone", "Mother Phone"),
    ("guardian_name", "Guardian Name"),
    ("guardian_relation", "Guardian Relation"),
    ("guardian_phone", "Guardian Phone"),
    ("annual_income", "Annual Income"),
    # Educational Qualifications
    ("tenth_board", "10th Board"),
    ("tenth_year", "10th Year"),
    ("tenth_percentage", "10th Percentage"),
    ("tenth_school", "10th School"),
    ("twelfth_board", "12th Board"),
    ("twelfth_year", "12th Year"),
    ("twelfth_percentage", "12th Percentage"),
    ("twelfth_school", "12th School"),
    ("previous_qualification", "Previous Qualification"),
    ("graduation_details", "Graduation Details"),
    # Course Application Details
    ("course_applied", "Course Applied"),
    ("application_number", "Application Number"),
    ("enrollment_number", "Enrollment Number"),
    ("date_of_admission", "Admission Date"),
]

# Fields to exclude from Excel exports (but keep in CSV/JSON for completeness)
EXCEL_EXCLUDE_FIELDS = {"ocr_provider", "additional_info"}

def form_to_csv_row(form: AdmissionForm) -> list[str]:
    row: list[str] = []
    for attr, _ in EXPORT_FIELDS:
        value = getattr(form, attr, None)
        if isinstance(value, datetime):
            value = value.isoformat()
        elif isinstance(value, FormStatus):
            value = value.value
        elif isinstance(value, dict):
            value = json.dumps(value, ensure_ascii=False)
        elif value is None:
            value = ""
        row.append(str(value) if not isinstance(value, str) else value)
    return row

def form_to_json_dict(form: AdmissionForm) -> Dict[str, Any]:
    record: Dict[str, Any] = {}
    for attr, _ in EXPORT_FIELDS:
        value = getattr(form, attr, None)
        if isinstance(value, datetime):
            record[attr] = value.isoformat()
        elif isinstance(value, FormStatus):
            record[attr] = value.value
        elif isinstance(value, dict):
            record[attr] = value
        else:
            record[attr] = value
    if record.get("additional_info") is None:
        record["additional_info"] = {}
    return record

from fastapi import Response
from fastapi.responses import StreamingResponse
from typing import Iterable

def export_to_excel_forms(forms: Iterable[AdmissionForm]) -> StreamingResponse:
    """Export forms to Excel format"""
    import pandas as pd
    
    # Convert iterator to list to check count and allow multiple iterations
    forms_list = list(forms) if not isinstance(forms, list) else forms
    logger.info(f"Exporting {len(forms_list)} forms to Excel")
    
    if not forms_list:
        # Return empty Excel file with headers
        logger.warning("No forms found for Excel export, returning empty file with headers")
        output = io.BytesIO()
        excel_fields = [(attr, header) for attr, header in EXPORT_FIELDS if attr not in EXCEL_EXCLUDE_FIELDS]
        headers = [header for _, header in excel_fields]
        df = pd.DataFrame(columns=headers)
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            df.to_excel(writer, sheet_name='Admission Forms', index=False)
            worksheet = writer.sheets['Admission Forms']
            # Format empty file with headers
            header_font = Font(name='Calibri', size=11, bold=True, color='FFFFFF')
            header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
            header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
            for idx in range(1, len(headers) + 1):
                cell = worksheet.cell(row=1, column=idx)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment
                cell.border = border
                worksheet.column_dimensions[get_column_letter(idx)].width = 15
            worksheet.row_dimensions[1].height = 25
        output.seek(0)
        return StreamingResponse(
            output,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": "attachment; filename=admission_forms.xlsx"}
        )
    
    # Filter out excluded fields for Excel
    excel_fields = [(attr, header) for attr, header in EXPORT_FIELDS if attr not in EXCEL_EXCLUDE_FIELDS]
    
    all_data = []
    for form in forms_list:
        try:
            record = form_to_json_dict(form)
            # Remove excluded fields
            for field in EXCEL_EXCLUDE_FIELDS:
                record.pop(field, None)
            
            # Log first form for debugging
            if len(all_data) == 0:
                logger.info(f"First form data sample - Form ID: {form.id}, Keys: {list(record.keys())[:10]}...")
            
            all_data.append(record)
        except Exception as e:
            logger.error(f"Error processing form {form.id if hasattr(form, 'id') else 'unknown'} for export: {e}", exc_info=True)
            # Continue with other forms even if one fails
    
    logger.info(f"Collected data for {len(all_data)} forms")
    
    if not all_data:
        # Return empty Excel file with headers
        logger.warning("No data collected for Excel export, returning empty file with headers")
        output = io.BytesIO()
        headers = [header for _, header in excel_fields]
        df = pd.DataFrame(columns=headers)
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            df.to_excel(writer, sheet_name='Admission Forms', index=False)
            worksheet = writer.sheets['Admission Forms']
            # Format empty file with headers
            header_font = Font(name='Calibri', size=11, bold=True, color='FFFFFF')
            header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
            header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
            for idx in range(1, len(headers) + 1):
                cell = worksheet.cell(row=1, column=idx)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment
                cell.border = border
                worksheet.column_dimensions[get_column_letter(idx)].width = 15
            worksheet.row_dimensions[1].height = 25
        output.seek(0)
        return StreamingResponse(
            output,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": "attachment; filename=admission_forms.xlsx"}
        )
    
    # Create DataFrame
    try:
        if not all_data:
            raise ValueError("No data to create DataFrame")
        
        df = pd.DataFrame(all_data)
        logger.info(f"Created DataFrame with {len(df)} rows and {len(df.columns)} columns")
        
        # Check if DataFrame is actually empty (no rows or no columns)
        if df.empty:
            logger.warning("DataFrame is empty after creation - checking data...")
            if all_data:
                logger.info(f"First record sample: {list(all_data[0].keys())[:10] if all_data[0] else 'empty dict'}")
                # Try to create DataFrame with explicit column order
                headers = [attr for attr, _ in excel_fields]
                df = pd.DataFrame(all_data, columns=headers)
                logger.info(f"Recreated DataFrame with explicit columns: {len(df)} rows, {len(df.columns)} columns")
            else:
                headers = [header for _, header in excel_fields]
                df = pd.DataFrame(columns=headers)
        else:
            # Ensure all expected columns exist (fill with None if missing)
            expected_attrs = [attr for attr, _ in excel_fields]
            for attr in expected_attrs:
                if attr not in df.columns:
                    df[attr] = None
                    logger.debug(f"Added missing column: {attr}")
    except Exception as e:
        logger.error(f"Error creating DataFrame: {e}", exc_info=True)
        # Fallback: create DataFrame with headers only
        headers = [header for _, header in excel_fields]
        df = pd.DataFrame(columns=headers)
        logger.warning("Created fallback DataFrame with headers only")
    
    # Map internal IDs to Headers (only for fields in excel_fields)
    header_map = {attr: header for attr, header in excel_fields}
    # Only rename columns that exist in the dataframe
    df = df.rename(columns={k: v for k, v in header_map.items() if k in df.columns})
    
    logger.info(f"Final DataFrame shape: {df.shape} (rows x columns), columns: {list(df.columns)[:10]}...")
    
    output = io.BytesIO()
    try:
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
        
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            df.to_excel(writer, sheet_name='Admission Forms', index=False)
            
            # Get the worksheet to format it
            worksheet = writer.sheets['Admission Forms']
            
            # Define styles
            header_font = Font(name='Calibri', size=11, bold=True, color='FFFFFF')
            header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
            header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            cell_font = Font(name='Calibri', size=10)
            cell_alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
            border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            
            # Format header row (row 1)
            for idx, col in enumerate(df.columns, start=1):
                cell = worksheet.cell(row=1, column=idx)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment
                cell.border = border
            
            # Format data rows
            for row_idx in range(2, len(df) + 2):  # Start from row 2 (after header)
                for col_idx in range(1, len(df.columns) + 1):
                    cell = worksheet.cell(row=row_idx, column=col_idx)
                    cell.font = cell_font
                    cell.alignment = cell_alignment
                    cell.border = border
            
            # Auto-adjust column widths with better calculation
            for idx, col in enumerate(df.columns, start=1):
                try:
                    # Calculate max length in column data
                    if len(df) > 0:
                        max_data_length = df[col].astype(str).map(len).max()
                    else:
                        max_data_length = 0
                    
                    # Get header length
                    header_length = len(str(col))
                    
                    # Calculate width: max of data and header, with padding
                    max_length = max(max_data_length, header_length)
                    
                    # Set width with better logic:
                    # - Minimum width: 10
                    # - Maximum width: 60 (increased from 50)
                    # - Add padding: 3 characters
                    # - For very short columns, use a reasonable minimum
                    if max_length < 10:
                        adjusted_width = 12
                    elif max_length > 57:
                        adjusted_width = 60
                    else:
                        adjusted_width = max_length + 3
                    
                    worksheet.column_dimensions[get_column_letter(idx)].width = adjusted_width
                except Exception as e:
                    # If column formatting fails, continue with other columns
                    logger.warning(f"Error formatting column {col}: {e}")
            
            # Set row heights
            worksheet.row_dimensions[1].height = 25  # Header row height
            for row_idx in range(2, len(df) + 2):
                worksheet.row_dimensions[row_idx].height = 18  # Data row height
            
            # Freeze header row
            worksheet.freeze_panes = 'A2'
            
            # Enable filter on header row
            worksheet.auto_filter.ref = worksheet.dimensions
        
        output.seek(0)
        file_size = len(output.getvalue())
        logger.info(f"Excel file created successfully, size: {file_size} bytes, rows: {len(df)}, columns: {len(df.columns)}")
        
        if file_size < 1000:  # Very small file might be empty
            logger.warning(f"Excel file is very small ({file_size} bytes), might be empty")
    except Exception as e:
        logger.error(f"Error creating Excel file: {e}", exc_info=True)
        raise HTTPException(status_code=500, detail=f"Error creating Excel file: {str(e)}")
    
    output.seek(0)
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=admission_forms.xlsx"}
    )

def export_to_pdf_forms(forms: Iterable[AdmissionForm]) -> StreamingResponse:
    """Export forms to PDF format"""
    try:
        from reportlab.lib import colors
        from reportlab.lib.pagesizes import A4, landscape
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.lib.units import inch
        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
        from reportlab.lib.enums import TA_CENTER
        
        output = io.BytesIO()
        doc = SimpleDocTemplate(output, pagesize=landscape(A4), topMargin=0.5*inch, bottomMargin=0.5*inch)
        
        styles = getSampleStyleSheet()
        title_style = ParagraphStyle('Title', parent=styles['Heading1'], alignment=TA_CENTER)
        
        elements = [Paragraph("Admission Forms Export", title_style), Spacer(1, 20)]
        
        # Select key fields for PDF (A4 landscape can't fit all 75 fields)
        key_fields = [
            ("id", "ID"), ("student_name", "Student Name"), ("status", "Status"),
            ("course_applied", "Course"), ("phone_number", "Phone"), ("upload_date", "Date")
        ]
        
        table_data = [[header for _, header in key_fields]]
        for form in forms:
            row = []
            for attr, _ in key_fields:
                val = getattr(form, attr, "")
                if isinstance(val, datetime): val = val.strftime("%Y-%m-%d")
                elif isinstance(val, FormStatus): val = val.value
                row.append(str(val) if val is not None else "-")
            table_data.append(row)
            
        t = Table(table_data)
        t.setStyle(TableStyle([
            ('BACKGROUND', (0,0), (-1,0), colors.navy),
            ('TEXTCOLOR', (0,0), (-1,0), colors.whitesmoke),
            ('ALIGN', (0,0), (-1,-1), 'CENTER'),
            ('FONTNAME', (0,0), (-1,0), 'Helvetica-Bold'),
            ('GRID', (0,0), (-1,-1), 0.5, colors.grey)
        ]))
        elements.append(t)
        
        doc.build(elements)
        output.seek(0)
        return StreamingResponse(
            output,
            media_type="application/pdf",
            headers={"Content-Disposition": "attachment; filename=admission_forms.pdf"}
        )
    except ImportError:
        raise HTTPException(status_code=500, detail="PDF export requires reportlab library. Install with: pip install reportlab")

def export_to_csv(forms: Iterable[AdmissionForm]) -> StreamingResponse:
    """Export forms to CSV format using a streaming response."""

    def row_iterator():
        output = io.StringIO()
        writer = csv.writer(output)
        writer.writerow([header for _, header in EXPORT_FIELDS])
        yield output.getvalue()
        output.seek(0)
        output.truncate(0)

        for form in forms:
            writer.writerow(form_to_csv_row(form))
            yield output.getvalue()
            output.seek(0)
            output.truncate(0)

    return StreamingResponse(
        row_iterator(),
        media_type="text/csv",
        headers={"Content-Disposition": "attachment; filename=admission_forms.csv"},
    )

def export_to_json_response(forms: Iterable[AdmissionForm], filters: Dict[str, Any]) -> Response:
    """Export forms to JSON format."""
    forms_list = list(forms)
    payload = {
        "generated_at": datetime.utcnow().isoformat(),
        "count": len(forms_list),
        "filters": filters,
        "forms": [form_to_json_dict(form) for form in forms_list],
    }

    return Response(
        content=json.dumps(payload, indent=2, ensure_ascii=False),
        media_type="application/json",
        headers={"Content-Disposition": "attachment; filename=admission_forms.json"},
    )

@router.get("/export")
async def export_forms(
    format: str = Query("csv", regex="^(csv|json|excel|pdf)$"),
    status: Optional[FormStatus] = Query(None),
    student_name: Optional[str] = Query(None),
    phone_number: Optional[str] = Query(None),
    email: Optional[str] = Query(None),
    enrollment_number: Optional[str] = Query(None),
    application_number: Optional[str] = Query(None),
    course_applied: Optional[str] = Query(None),
    date_from: Optional[datetime] = Query(None),
    date_to: Optional[datetime] = Query(None),
    form_ids: Optional[str] = Query(None, description="Comma-separated list of form IDs to export"),
    db: Session = Depends(get_db),
    user: CurrentUser = Depends(RequireStaffOrAdmin),
):
    """Export forms to CSV, JSON, Excel or PDF"""
    query = db.query(AdmissionForm)

    # If form_ids are provided, filter by those IDs first
    if form_ids:
        try:
            ids_list = [int(id.strip()) for id in form_ids.split(',') if id.strip().isdigit()]
            if ids_list:
                query = query.filter(AdmissionForm.id.in_(ids_list))
        except ValueError:
            # Invalid form_ids format, ignore it
            pass

    query = apply_form_filters(
        query,
        student_name=student_name,
        phone_number=phone_number,
        email=email,
        enrollment_number=enrollment_number,
        application_number=application_number,
        course_applied=course_applied,
        status=status,
        date_from=date_from,
        date_to=date_to,
    )

    # If no status filter is provided, default to VERIFIED forms only for safety
    # unless some other identifying search parameter is provided or form_ids are specified
    if status is None and not any([student_name, phone_number, email, enrollment_number, application_number, form_ids]):
        query = query.filter(AdmissionForm.status == FormStatus.VERIFIED)
    
    forms = query.order_by(AdmissionForm.upload_date.desc(), AdmissionForm.id.desc()).all()

    filters_snapshot = {
        "student_name": student_name,
        "phone_number": phone_number,
        "email": email,
        "enrollment_number": enrollment_number,
        "application_number": application_number,
        "course_applied": course_applied,
        "status": status.value if status else None,
        "date_from": date_from.isoformat() if date_from else None,
        "date_to": date_to.isoformat() if date_to else None,
    }
    filters_active = {key: value for key, value in filters_snapshot.items() if value}

    logger.info(
        "Exporting forms format=%s count=%s filters=%s form_ids=%s",
        format,
        len(forms),
        filters_active,
        form_ids,
    )
    
    # Convert to list to allow multiple iterations and better error handling
    forms_list = list(forms)
    logger.info(f"Converted forms query to list: {len(forms_list)} forms")
    
    if format == "csv":
        return export_to_csv(forms_list)
    elif format == "excel":
        return export_to_excel_forms(forms_list)
    elif format == "pdf":
        return export_to_pdf_forms(forms_list)
    else:
        return export_to_json_response(forms_list, filters_active)

# --- End Export Related Logic ---

@router.get("/", response_model=List[FormDetailResponse])
async def list_forms(
    skip: int = Query(0, ge=0),
    limit: int = Query(20, ge=1, le=50000),  # Increased limit for "select all" functionality
    status: Optional[FormStatus] = None,
    sort_by: Optional[str] = Query("upload_date", description="Field to sort by"),
    sort_order: Optional[str] = Query("desc", regex="^(asc|desc)$"),
    db: Session = Depends(get_db),
    user: CurrentUser = Depends(RequireAnyAuth),
):
    """List all admission forms with pagination and sorting"""
    from sqlalchemy.orm import selectinload
    from backend.models.document import DocumentResponse
    
    # Use selectinload to eagerly load documents in a separate efficient query
    # This avoids N+1 queries and cartesian product issues from joinedload
    query = db.query(AdmissionForm).options(selectinload(AdmissionForm.documents))
    
    if status:
        query = query.filter(AdmissionForm.status == status)
    
    # Map sort_by field - validate it exists on the model
    try:
        sort_field = getattr(AdmissionForm, sort_by, AdmissionForm.upload_date)
    except AttributeError:
        sort_field = AdmissionForm.upload_date
    
    if sort_order == "desc":
        query = query.order_by(desc(sort_field))
    else:
        query = query.order_by(asc(sort_field))
    
    # Execute query with limit/offset
    forms = query.offset(skip).limit(limit).all()
    
    # Build response - documents are already loaded via selectinload
    result = []
    for form in forms:
        form_data = FormDetailResponse.model_validate(form)
        # Documents are already loaded via relationship, just convert to response model
        form_data.documents = [DocumentResponse.model_validate(doc) for doc in form.documents]
        result.append(form_data)
    
    logger.debug(f"Listed {len(result)} forms (skip={skip}, limit={limit}, status={status})")
    return result

@router.get("/ids", response_model=List[int])
async def list_form_ids(
    status: Optional[FormStatus] = None,
    sort_by: Optional[str] = Query("upload_date", description="Field to sort by"),
    sort_order: Optional[str] = Query("desc", regex="^(asc|desc)$"),
    db: Session = Depends(get_db),
    user: CurrentUser = Depends(RequireAnyAuth),
):
    """Get all form IDs matching filters (lightweight endpoint for select all)"""
    query = db.query(AdmissionForm.id)
    
    if status:
        query = query.filter(AdmissionForm.status == status)
    
    # Map sort_by field - validate it exists on the model
    try:
        sort_field = getattr(AdmissionForm, sort_by, AdmissionForm.upload_date)
    except AttributeError:
        sort_field = AdmissionForm.upload_date
    
    if sort_order == "desc":
        query = query.order_by(desc(sort_field))
    else:
        query = query.order_by(asc(sort_field))
    
    # Execute query and return just IDs
    form_ids = [row[0] for row in query.all()]
    
    logger.debug(f"Listed {len(form_ids)} form IDs (status={status})")
    return form_ids

@router.get("/search/results", response_model=List[FormResponse])
async def search_forms(
    skip: int = Query(0, ge=0),
    limit: int = Query(100, ge=1, le=1000),
    status: Optional[FormStatus] = Query(None),
    student_name: Optional[str] = Query(None),
    phone_number: Optional[str] = Query(None),
    email: Optional[str] = Query(None),
    enrollment_number: Optional[str] = Query(None),
    application_number: Optional[str] = Query(None),
    course_applied: Optional[str] = Query(None),
    date_from: Optional[datetime] = Query(None),
    date_to: Optional[datetime] = Query(None),
    db: Session = Depends(get_db),
    user: CurrentUser = Depends(RequireAnyAuth),
):
    """
    Search admission forms with filters.
    Alias for list_forms to match frontend expectation.
    """
    return await list_forms(
        skip=skip, limit=limit, status=status,
        student_name=student_name, phone_number=phone_number,
        email=email, enrollment_number=enrollment_number,
        application_number=application_number, course_applied=course_applied,
        date_from=date_from, date_to=date_to, db=db, user=user
    )

@router.get("/{form_id}", response_model=FormDetailResponse)
async def get_form(
    form_id: int, db: Session = Depends(get_db),
    user: CurrentUser = Depends(RequireAnyAuth),
):
    """Get detailed information about a specific form"""
    form = db.query(AdmissionForm).filter(AdmissionForm.id == form_id).first()
    if not form:
        raise HTTPException(status_code=404, detail="Form not found")
    
    # Get associated documents
    documents = db.query(StudentDocument).filter(
        StudentDocument.form_id == form_id
    ).order_by(StudentDocument.upload_date.desc()).all()
    
    from backend.models.document import DocumentResponse
    form_data = FormDetailResponse.model_validate(form)
    form_data.documents = [DocumentResponse.model_validate(doc) for doc in documents]
    
    return form_data

@router.post(
    "/{form_id}/extract",
    response_model=FormExtractionResponse,
    summary="Re-extract a form using the selected OCR provider",
)
async def re_extract_form(
    form_id: int,
    ocr_provider: Optional[str] = None,
    db: Session = Depends(get_db),
    user: CurrentUser = Depends(RequireStaffOrAdmin),
):
    """Re-extract text from a form using a different or same OCR provider"""
    form = db.query(AdmissionForm).filter(AdmissionForm.id == form_id).first()
    if not form:
        raise HTTPException(status_code=404, detail="Form not found")
    
    try:
        from backend.utils.extraction_pipeline import run_enhanced_extraction
        
        # Use the unified enhanced extraction pipeline
        ocr_result = await run_enhanced_extraction(form, db, ocr_provider)
        
        logger.info(
            "Re-extracted form %s with provider %s (pages=%s, confidence=%s)",
            form_id,
            ocr_result.get("provider", "unknown"),
            ocr_result.get("pages_processed", 0),
            ocr_result.get("confidence", 0),
        )
        
        return FormExtractionResponse(
            message="Re-extraction completed",
            result=ExtractedData(**ocr_result),
        )
        
    except Exception as e:
        form.status = FormStatus.ERROR
        db.commit()
        # logger.exception("Re-extraction failed") # extraction_pipeline logs internal errors
        raise HTTPException(status_code=500, detail=f"Re-extraction failed: {str(e)}")

@router.put("/{form_id}/verify", response_model=FormDetailResponse)
async def verify_form(
    form_id: int,
    verification: FormVerification,
    db: Session = Depends(get_db),
    user: CurrentUser = Depends(RequireStaffOrAdmin),
):
    """Save verified/corrected student information"""
    form = db.query(AdmissionForm).filter(AdmissionForm.id == form_id).first()
    if not form:
        raise HTTPException(status_code=404, detail="Form not found")
    
    # Update form with ALL verified data fields
    all_fields = [
        # Basic Personal Details
        'student_name', 'first_name', 'middle_name', 'surname',
        'date_of_birth', 'gender', 'category', 'nationality', 'religion',
        'aadhar_number', 'blood_group', 'below_poverty_line', 'minority_category',
        
        # Academic & Admission Details
        'academic_session', 'course', 'admission_category', 'admission_category_other',
        'du_portal_form_number', 'cuet_score', 'college_roll_no', 'date_of_admission',
        'course_applied', 'application_number', 'enrollment_number', 'date_of_admission',
        'du_enrollment_number', 'hindi_medium_preference',
        
        # Address Details
        'permanent_address', 'permanent_address_line1', 'permanent_address_line2', 
        'permanent_address_line3', 'permanent_state', 'permanent_pincode',
        'correspondence_address', 'correspondence_address_line1', 'correspondence_address_line2',
        'correspondence_address_line3', 'correspondence_state', 'correspondence_pincode',
        'pincode', 'city', 'state',
        
        # Contact Details
        'phone_number', 'alternate_phone', 'email', 
        'emergency_contact_name', 'emergency_contact_phone',
        
        # Mother's Details
        'mother_name', 'mother_occupation', 'mother_designation', 'mother_organization',
        'mother_email', 'mother_mobile', 'mother_landline_code', 'mother_landline', 'mother_phone',
        
        # Father's Details
        'father_name', 'father_occupation', 'father_designation', 'father_organization',
        'father_email', 'father_mobile', 'father_landline_code', 'father_landline', 'father_phone',
        
        # Guardian Details
        'guardian_name', 'guardian_relation', 'guardian_residential_address', 'guardian_organization',
        'guardian_email', 'guardian_mobile', 'guardian_landline_code', 'guardian_landline', 'guardian_phone',
        
        # Family Income
        'annual_income',
        
        # Academic History
        'tenth_board', 'tenth_year', 'tenth_percentage', 'tenth_school',
        'twelfth_board', 'twelfth_year', 'twelfth_percentage', 'twelfth_school',
        'twelfth_roll_number', 'twelfth_institution', 'hindi_studied_upto',
        'previous_qualification', 'graduation_details',
        
        # Certificate Details
        'category_certificate_authority', 'category_certificate_number', 'category_certificate_date',
        'disability_percentage', 'disability_type', 'udid_number',
        
        # CUET Marks
        'cuet_subject_1', 'cuet_total_score_1', 'cuet_score_obtained_1',
        'cuet_subject_2', 'cuet_total_score_2', 'cuet_score_obtained_2',
        'cuet_subject_3', 'cuet_total_score_3', 'cuet_score_obtained_3',
        'cuet_subject_4', 'cuet_total_score_4', 'cuet_score_obtained_4',
        'cuet_subject_5', 'cuet_total_score_5', 'cuet_score_obtained_5',
        'cuet_subject_6', 'cuet_total_score_6', 'cuet_score_obtained_6',
        'cuet_total_score',
        
        # Document Checklist
        'doc_admission_form', 'doc_undertaking_ragging', 'doc_photographs',
        'doc_cuet_scorecard', 'doc_class_xii_marksheet', 'doc_class_x_certificate',
        'doc_class_xii_certificate', 'doc_character_certificate', 'doc_transfer_certificate',
        'doc_hindi_certificate', 'doc_caste_certificate', 'doc_sports_eca',
        'doc_originals', 'doc_photo_id',
    ]
    
    for field in all_fields:
        value = getattr(verification, field, None)
        if value is not None and hasattr(form, field):
            setattr(form, field, value)
    
    # Sync category and admission_category (they are the same field)
    if form.admission_category and not form.category:
        form.category = form.admission_category
    elif form.category and not form.admission_category:
        form.admission_category = form.category
    elif form.admission_category:
        form.category = form.admission_category

    form.additional_info = verification.additional_info or {}

    # Validate required field: student_name
    if not verification.student_name or not verification.student_name.strip():
        raise HTTPException(
            status_code=400,
            detail="Student name is required. A form cannot be verified without a student name."
        )
    
    # Create or update student profile using verified data
    try:
        profile = get_or_create_student_profile(
            db, 
            student_name=verification.student_name,
            aadhar_number=verification.aadhar_number,
            roll_number=verification.college_roll_no or form.college_roll_no
        )
        
        # Update profile with verification flag
        profile.is_verified = True
        if not profile.roll_number and (verification.college_roll_no or form.college_roll_no):
            profile.roll_number = verification.college_roll_no or form.college_roll_no
        
        db.flush()
        
        # Link form to profile
        form.student_profile_id = profile.id
    except Exception as e:
        # Critical error: Profile creation failed. Log full traceback.
        logger.error(f"Failed to create/link student profile for form {form.id}: {e}", exc_info=True)
        raise HTTPException(
            status_code=500,
            detail=f"Failed to create student profile: {str(e)}"
        )
    
    form.status = FormStatus.VERIFIED
    form.verified_date = datetime.utcnow()
    
    # Rename file to include student name (only if not already renamed)
    try:
        from backend.utils.file_handler import rename_form_file, sanitize_filename
        
        # Check if filename already contains student name (avoid renaming multiple times)
        sanitized_name = sanitize_filename(verification.student_name).lower()
        current_filename_lower = form.filename.lower()
        
        # Only rename if filename doesn't already start with the student name pattern
        # Check if filename starts with sanitized name followed by underscore and form_id
        expected_prefix = f"{sanitized_name}_{form.id}".lower()
        
        if sanitized_name and not current_filename_lower.startswith(expected_prefix):
            new_file_path, new_filename = rename_form_file(
                form.file_path,
                verification.student_name,
                form.id
            )
            form.file_path = new_file_path
            form.filename = new_filename
            logger.info(f"Renamed form {form.id} file to: {new_filename}")
    except Exception as e:
        # Log error but don't fail the verification
        logger.warning(f"Could not rename form file: {e}")
    
    # Track corrections for continuous improvement
    try:
        from backend.utils.continuous_improvement import ContinuousImprovementManager
        
        improvement_manager = ContinuousImprovementManager()
        
        # Compare verified values with original extracted values to find corrections
        extracted_data = form.extracted_data or {}
        structured_data = extracted_data.get('structured_data', {})
        
        # Track corrections for each field
        for field_name in ['student_name', 'date_of_birth', 'gender', 'category', 
                          'nationality', 'religion', 'aadhar_number', 'blood_group',
                          'permanent_address', 'correspondence_address', 'pincode', 
                          'city', 'state', 'phone_number', 'alternate_phone', 'email',
                          'emergency_contact_name', 'emergency_contact_phone',
                          'father_name', 'father_occupation', 'father_phone',
                          'mother_name', 'mother_occupation', 'mother_phone',
                          'guardian_name', 'guardian_relation', 'guardian_phone',
                          'annual_income', 'tenth_board', 'tenth_year', 'tenth_percentage',
                          'tenth_school', 'twelfth_board', 'twelfth_year', 
                          'twelfth_percentage', 'twelfth_school', 'previous_qualification',
                          'graduation_details', 'course_applied', 'application_number',
                          'enrollment_number', 'date_of_admission']:
            
            verified_value = getattr(verification, field_name, None)
            original_value = structured_data.get(field_name) or getattr(form, field_name, None)
            
            # Record correction if values differ
            if verified_value and original_value and str(verified_value).strip() != str(original_value).strip():
                confidence = extracted_data.get('confidence', 0) / 100.0 if extracted_data.get('confidence') else None
                improvement_manager.record_correction(
                    form_id=form_id,
                    field_name=field_name,
                    original_value=str(original_value),
                    corrected_value=str(verified_value),
                    confidence=confidence
                )
    except Exception as e:
        # Log but don't fail verification if improvement tracking fails
        logger.warning(f"Failed to track corrections for continuous improvement: {e}")
    
    # Train Google OCR with verified corrections
    try:
        from backend.training.train_google_ocr import GoogleOCRTrainer
        
        trainer = GoogleOCRTrainer()
        
        # Get raw OCR text from extracted data
        extracted_data = form.extracted_data or {}
        raw_text = extracted_data.get('raw_text', '')
        
        if raw_text:
            # Get extracted fields (before verification)
            extracted_fields = {}
            structured_data = extracted_data.get('structured_data', {})
            for key, value in structured_data.items():
                if isinstance(value, str) and value:
                    extracted_fields[key] = value
            
            # Get verified fields (after verification)
            verified_fields = {}
            for field_name in ['student_name', 'date_of_birth', 'gender', 'category',
                              'nationality', 'religion', 'aadhar_number', 'blood_group',
                              'permanent_address', 'correspondence_address', 'pincode',
                              'phone_number', 'email', 'father_name', 'mother_name',
                              'course_applied', 'application_number', 'enrollment_number']:
                value = getattr(verification, field_name, None)
                if value:
                    verified_fields[field_name] = str(value)
            
            if verified_fields:
                # Add to training data
                result = trainer.add_verified_sample(
                    form_id=str(form_id),
                    raw_ocr_text=raw_text,
                    extracted_fields=extracted_fields,
                    verified_fields=verified_fields,
                    image_path=form.file_path
                )
                logger.info(f"Added form {form_id} to Google OCR training: accuracy={result.get('accuracy', 0):.2f}")
    except Exception as e:
        logger.warning(f"Failed to add form to Google OCR training: {e}")
    
    # Record verification in unified training manager (V2)
    try:
        from backend.utils.training_manager import training_manager
        
        extracted_data = form.extracted_data or {}
        original_data = extracted_data.get('structured_data', {})
        verified_data = {}
        
        for field_name in all_fields:
            value = getattr(verification, field_name, None)
            if value:
                verified_data[field_name] = str(value)
        
        training_manager.record_verification(
            form_id=form_id,
            original_data=original_data,
            verified_data=verified_data,
            raw_ocr_text=extracted_data.get('raw_text')
        )
        logger.info(f"Recorded verification for form {form_id} in training manager")
    except Exception as e:
        logger.warning(f"Failed to record in training manager: {e}")
    
    # Automatically create annotation from verified data for training
    try:
        from backend.api.routes.annotation import AnnotationField, AnnotationCheckbox
        
        annotation_fields = []
        # Create annotation fields from verified data
        field_mapping = {
            'student_name': 'student_name',
            'date_of_birth': 'date_of_birth',
            'gender': 'gender',
            'category': 'category',
            'nationality': 'nationality',
            'religion': 'religion',
            'aadhar_number': 'aadhar_number',
            'blood_group': 'blood_group',
            'permanent_address': 'permanent_address',
            'correspondence_address': 'correspondence_address',
            'pincode': 'pincode',
            'city': 'city',
            'state': 'state',
            'phone_number': 'phone_number',
            'alternate_phone': 'alternate_phone',
            'email': 'email',
            'emergency_contact_name': 'emergency_contact_name',
            'emergency_contact_phone': 'emergency_contact_phone',
            'father_name': 'father_name',
            'father_occupation': 'father_occupation',
            'father_phone': 'father_phone',
            'mother_name': 'mother_name',
            'mother_occupation': 'mother_occupation',
            'mother_phone': 'mother_phone',
            'guardian_name': 'guardian_name',
            'guardian_relation': 'guardian_relation',
            'guardian_phone': 'guardian_phone',
            'annual_income': 'annual_income',
            'tenth_board': 'tenth_board',
            'tenth_year': 'tenth_year',
            'tenth_percentage': 'tenth_percentage',
            'tenth_school': 'tenth_school',
            'twelfth_board': 'twelfth_board',
            'twelfth_year': 'twelfth_year',
            'twelfth_percentage': 'twelfth_percentage',
            'twelfth_school': 'twelfth_school',
            'previous_qualification': 'previous_qualification',
            'graduation_details': 'graduation_details',
            'course_applied': 'course_applied',
            'application_number': 'application_number',
            'enrollment_number': 'enrollment_number',
            'admission_date': 'date_of_admission',
        }
        
        for field_key, field_name in field_mapping.items():
            value = getattr(verification, field_key, None)
            if value and str(value).strip():
                annotation_fields.append(AnnotationField(
                    field_name=field_name,
                    value=str(value).strip(),
                    page_number=1,
                    confidence=1.0  # Verified data has 100% confidence
                ))
        
        # Create key-value pairs for training
        key_value_pairs = {f.field_name: f.value for f in annotation_fields}
        
        # Store annotation in additional_info
        form.additional_info['annotation'] = {
            'fields': [f.dict() for f in annotation_fields],
            'checkboxes': [],
            'key_value_pairs': key_value_pairs,
            'notes': 'Auto-created from verified form data',
            'annotated_at': datetime.utcnow().isoformat(),
            'annotated_by': 'verification-api'
        }
        
        logger.info(f"Created annotation for form {form_id} with {len(annotation_fields)} fields")
    except Exception as e:
        # Log error but don't fail verification
        logger.warning(f"Failed to create annotation for form {form_id}: {e}")
    
    db.commit()
    db.refresh(form)
    
    # Get associated documents
    documents = db.query(StudentDocument).filter(
        StudentDocument.form_id == form.id
    ).order_by(StudentDocument.upload_date.desc()).all()
    
    from backend.models.document import DocumentResponse
    form_data = FormDetailResponse.model_validate(form)
    form_data.documents = [DocumentResponse.model_validate(doc) for doc in documents]
    
    return form_data

@router.put("/{form_id}", response_model=FormDetailResponse)
async def update_form(
    form_id: int,
    verification: FormVerification,
    verify: bool = Query(False, description="If True, mark form as verified and link to student profile"),
    db: Session = Depends(get_db),
    user: CurrentUser = Depends(RequireStaffOrAdmin),
):
    """Update form data. Use verify=True to mark as verified and link to student profile."""
    if verify:
        return await verify_form(form_id, verification, db, user)
    
    form = db.query(AdmissionForm).filter(AdmissionForm.id == form_id).first()
    if not form:
        raise HTTPException(status_code=404, detail="Form not found")
        
    # Standard update logic
    for field, value in verification.model_dump(exclude_unset=True).items():
        if hasattr(form, field):
            setattr(form, field, value)
            
    db.commit()
    db.refresh(form)
    return form

def delete_form_files(form: AdmissionForm):
    """Delete all physical files associated with a form and its documents"""
    from pathlib import Path
    import os
    upload_dir = Path(settings.UPLOAD_DIR).resolve()
    
    # Delete form's own file
    try:
        if form.file_path:
            full_file_path = upload_dir / form.file_path
            if full_file_path.exists():
                os.remove(full_file_path)
    except Exception as e:
        logger.warning(f"Could not delete file for form {form.id}: {e}")
        
    # Delete associated documents' files
    for doc in form.documents:
        try:
            if doc.file_path:
                full_doc_path = upload_dir / doc.file_path
                if full_doc_path.exists():
                    os.remove(full_doc_path)
        except Exception as e:
            logger.warning(f"Could not delete file for document {doc.id}: {e}")

@router.delete("/{form_id}", status_code=204)
async def delete_form(
    form_id: int, db: Session = Depends(get_db),
    user: CurrentUser = Depends(RequireStaffOrAdmin),
):
    """Delete a specific admission form and its files"""
    form = db.query(AdmissionForm).filter(AdmissionForm.id == form_id).first()
    if not form:
        raise HTTPException(status_code=404, detail="Form not found")
        
    delete_form_files(form)
    db.delete(form)
    db.commit()
    return None

@router.post("/bulk-delete", status_code=204)
async def bulk_delete_forms(
    form_ids: List[int] = Body(...),
    db: Session = Depends(get_db),
    user: CurrentUser = Depends(RequireStaffOrAdmin),
):
    """Bulk delete admission forms"""
    forms = db.query(AdmissionForm).filter(AdmissionForm.id.in_(form_ids)).all()
    
    for form in forms:
        delete_form_files(form)
        db.delete(form)
        
    db.commit()
    return None
