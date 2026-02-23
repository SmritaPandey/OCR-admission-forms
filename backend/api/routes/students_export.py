"""
Student Export/Import Routes
Handles CSV and Excel export/import for student records
"""
from fastapi import APIRouter, HTTPException, Depends, Query, UploadFile, File
from sqlalchemy.orm import Session
from sqlalchemy import or_, and_, asc, desc, func
from typing import Optional, List
from backend.database import get_db, StudentProfile, AdmissionForm, StudentDocument
from backend.api.dependencies import RequireStaffOrAdmin
from backend.models.auth_models import CurrentUser
from datetime import datetime
import csv
import io
from typing import Dict, Any
import pandas as pd
import logging

router = APIRouter()
logger = logging.getLogger(__name__)


def get_excel_column_letter(col_idx: int) -> str:
    """
    Convert a column index (1-based) to Excel column letter(s).
    Examples: 1 -> A, 2 -> B, 26 -> Z, 27 -> AA, 28 -> AB, etc.
    """
    result = ""
    while col_idx > 0:
        col_idx -= 1
        result = chr(65 + (col_idx % 26)) + result
        col_idx //= 26
    return result

# Sort options
SORT_OPTIONS = {
    "name": "student_name",
    "roll_number": "roll_number",
    "aadhar": "aadhar_number",
    "created": "created_date",
    "updated": "updated_date",
}

def get_student_data_query(
    db: Session,
    student_name: Optional[str] = None,
    roll_number: Optional[str] = None,
    aadhar_number: Optional[str] = None,
    phone_number: Optional[str] = None,
    email: Optional[str] = None,
    course_applied: Optional[str] = None,
    academic_session: Optional[str] = None,
    gender: Optional[str] = None,
    category: Optional[str] = None,
    city: Optional[str] = None,
    state: Optional[str] = None,
    sort_by: str = "updated",
    sort_order: str = "desc",
    is_verified: bool = True  # Default to only verified students
):
    """Build query for student records with filters and sorting"""
    query = db.query(StudentProfile)
    
    # Filter by verified status (default to verified only)
    # If is_verified=True, include profiles that are verified OR have verified forms
    if is_verified is not None and is_verified:
        # Include profiles that are verified OR have at least one verified form
        from backend.database import FormStatus
        # Get profile IDs that have verified forms
        verified_form_profile_ids = [
            pid[0] for pid in db.query(AdmissionForm.student_profile_id).filter(
                AdmissionForm.status == FormStatus.VERIFIED,
                AdmissionForm.student_profile_id.isnot(None)
            ).distinct().all() if pid[0] is not None
        ]
        
        if verified_form_profile_ids:
            query = query.filter(
                or_(
                    StudentProfile.is_verified == True,
                    StudentProfile.id.in_(verified_form_profile_ids)
                )
            )
        else:
            # If no verified forms, only show verified profiles
            query = query.filter(StudentProfile.is_verified == True)
    elif is_verified is not None and not is_verified:
        query = query.filter(StudentProfile.is_verified == False)
    
    # Build filters
    profile_filters = []
    if student_name:
        profile_filters.append(StudentProfile.student_name.ilike(f"%{student_name}%"))
    if roll_number:
        profile_filters.append(StudentProfile.roll_number.ilike(f"%{roll_number}%"))
    if aadhar_number:
        profile_filters.append(StudentProfile.aadhar_number.ilike(f"%{aadhar_number}%"))
    
    # Join with forms if form filters are provided
    has_form_filters = any([phone_number, email, course_applied, academic_session, gender, category, city, state])
    if has_form_filters:
        query = query.join(AdmissionForm, StudentProfile.id == AdmissionForm.student_profile_id)
        form_filters = []
        if phone_number:
            form_filters.append(AdmissionForm.phone_number.ilike(f"%{phone_number}%"))
        if email:
            form_filters.append(AdmissionForm.email.ilike(f"%{email}%"))
        if course_applied:
            form_filters.append(AdmissionForm.course_applied.ilike(f"%{course_applied}%"))
        if academic_session:
            # Filter by academic session - can match full session string or just the year
            form_filters.append(
                or_(
                    AdmissionForm.academic_session.ilike(f"%{academic_session}%"),
                    AdmissionForm.academic_session.ilike(f"%{academic_session}-%"),
                    AdmissionForm.academic_session.ilike(f"%-{academic_session}%")
                )
            )
        if gender:
            # Use case-insensitive exact match for gender (not partial match)
            form_filters.append(func.lower(AdmissionForm.gender) == func.lower(gender))
        if category:
            form_filters.append(AdmissionForm.category.ilike(f"%{category}%"))
        if city:
            form_filters.append(AdmissionForm.city.ilike(f"%{city}%"))
        if state:
            form_filters.append(AdmissionForm.state.ilike(f"%{state}%"))
        if form_filters:
            query = query.filter(and_(*form_filters))
        query = query.distinct()
    
    if profile_filters:
        query = query.filter(and_(*profile_filters))
    
    # Apply sorting
    sort_column = SORT_OPTIONS.get(sort_by, "updated_date")
    if sort_order.lower() == "asc":
        query = query.order_by(asc(getattr(StudentProfile, sort_column)))
    else:
        query = query.order_by(desc(getattr(StudentProfile, sort_column)))
    
    return query

def get_student_export_data(profile: StudentProfile, db: Session) -> Dict[str, Any]:
    """Get comprehensive student data for export with ALL form fields"""
    # Get latest verified form, or most recent form
    latest_form = db.query(AdmissionForm).filter(
        AdmissionForm.student_profile_id == profile.id
    ).order_by(AdmissionForm.upload_date.desc()).first()
    
    # Helper function to safely get attribute value
    def safe_get_attr(obj, attr_name, default=""):
        """Safely get attribute value, returning default if attribute doesn't exist"""
        try:
            return getattr(obj, attr_name, default) or default
        except AttributeError:
            return default
    
    # Start with profile data
    data = {
        "ID": profile.id,
        "Student Name": profile.student_name or "",
        "Roll Number": profile.roll_number or "",
        "Aadhar Number": profile.aadhar_number or "",
        "Created Date": profile.created_date.strftime("%Y-%m-%d %H:%M:%S") if profile.created_date else "",
        "Updated Date": profile.updated_date.strftime("%Y-%m-%d %H:%M:%S") if profile.updated_date else "",
    }
    
    # Add ALL form fields if available
    if latest_form:
        # Helper function to safely get attribute value
        def safe_get_attr(attr_name, default=""):
            """Safely get attribute value, returning default if attribute doesn't exist"""
            try:
                value = getattr(latest_form, attr_name, None)
                return value or default
            except (AttributeError, TypeError):
                return default
        
        # Academic & Admission Details
        data.update({
            "Academic Session": safe_get_attr("academic_session"),
            "Course": safe_get_attr("course"),
            "Admission Category": safe_get_attr("admission_category"),
            "Admission Category Other": safe_get_attr("admission_category_other"),
            "DU Portal Form Number": safe_get_attr("du_portal_form_number"),
            "CUET Score": safe_get_attr("cuet_score"),
            "Total CUET Score": safe_get_attr("cuet_total_score"),
            "College Roll No": safe_get_attr("college_roll_no"),
            "Date of Admission": safe_get_attr("date_of_admission"),
        })
        
        # CUET Marks (all 6 subjects)
        data.update({
            "CUET Subject 1": safe_get_attr("cuet_subject_1"),
            "CUET Total Score 1": safe_get_attr("cuet_total_score_1"),
            "CUET Score Obtained 1": safe_get_attr("cuet_score_obtained_1"),
            "CUET Subject 2": safe_get_attr("cuet_subject_2"),
            "CUET Total Score 2": safe_get_attr("cuet_total_score_2"),
            "CUET Score Obtained 2": safe_get_attr("cuet_score_obtained_2"),
            "CUET Subject 3": safe_get_attr("cuet_subject_3"),
            "CUET Total Score 3": safe_get_attr("cuet_total_score_3"),
            "CUET Score Obtained 3": safe_get_attr("cuet_score_obtained_3"),
            "CUET Subject 4": safe_get_attr("cuet_subject_4"),
            "CUET Total Score 4": safe_get_attr("cuet_total_score_4"),
            "CUET Score Obtained 4": safe_get_attr("cuet_score_obtained_4"),
            "CUET Subject 5": safe_get_attr("cuet_subject_5"),
            "CUET Total Score 5": safe_get_attr("cuet_total_score_5"),
            "CUET Score Obtained 5": safe_get_attr("cuet_score_obtained_5"),
            "CUET Subject 6": safe_get_attr("cuet_subject_6"),
            "CUET Total Score 6": safe_get_attr("cuet_total_score_6"),
            "CUET Score Obtained 6": safe_get_attr("cuet_score_obtained_6"),
        })
        
        # Personal Details
        data.update({
            "First Name": safe_get_attr("first_name"),
            "Middle Name": safe_get_attr("middle_name"),
            "Surname": safe_get_attr("surname"),
            "Date of Birth": safe_get_attr("date_of_birth"),
            "Gender": safe_get_attr("gender"),
            "Category": safe_get_attr("category"),
            "Nationality": safe_get_attr("nationality"),
            "Religion": safe_get_attr("religion"),
            "Blood Group": safe_get_attr("blood_group"),
            "Below Poverty Line": safe_get_attr("below_poverty_line"),
            "Minority Category": safe_get_attr("minority_category"),
            "Annual Income": safe_get_attr("annual_income"),
        })
        
        # Permanent Address
        data.update({
            "Permanent Address Line 1": safe_get_attr("permanent_address_line1"),
            "Permanent Address Line 2": safe_get_attr("permanent_address_line2"),
            "Permanent Address Line 3": safe_get_attr("permanent_address_line3"),
            "Permanent State": safe_get_attr("permanent_state"),
            "Permanent Pincode": safe_get_attr("permanent_pincode"),
            "Permanent Address": safe_get_attr("permanent_address"),
        })
        
        # Correspondence Address
        data.update({
            "Correspondence Address Line 1": safe_get_attr("correspondence_address_line1"),
            "Correspondence Address Line 2": safe_get_attr("correspondence_address_line2"),
            "Correspondence Address Line 3": safe_get_attr("correspondence_address_line3"),
            "Correspondence State": safe_get_attr("correspondence_state"),
            "Correspondence Pincode": safe_get_attr("correspondence_pincode"),
            "Correspondence Address": safe_get_attr("correspondence_address"),
            "City": safe_get_attr("city"),
            "State": safe_get_attr("state"),
            "Pincode": safe_get_attr("pincode"),
        })
        
        # Contact Details
        data.update({
            "Email": safe_get_attr("email"),
            "Phone Number": safe_get_attr("phone_number"),
            "Alternate Phone": safe_get_attr("alternate_phone"),
            "Emergency Contact Name": safe_get_attr("emergency_contact_name"),
            "Emergency Contact Phone": safe_get_attr("emergency_contact_phone"),
        })
        
        # Mother's Details
        data.update({
            "Mother Name": safe_get_attr("mother_name"),
            "Mother Occupation": safe_get_attr("mother_occupation"),
            "Mother Designation": safe_get_attr("mother_designation"),
            "Mother Organization": safe_get_attr("mother_organization"),
            "Mother Email": safe_get_attr("mother_email"),
            "Mother Mobile": safe_get_attr("mother_mobile"),
            "Mother Landline Code": safe_get_attr("mother_landline_code"),
            "Mother Landline": safe_get_attr("mother_landline"),
            "Mother Phone": safe_get_attr("mother_phone"),
        })
        
        # Father's Details
        data.update({
            "Father Name": safe_get_attr("father_name"),
            "Father Occupation": safe_get_attr("father_occupation"),
            "Father Designation": safe_get_attr("father_designation"),
            "Father Organization": safe_get_attr("father_organization"),
            "Father Email": safe_get_attr("father_email"),
            "Father Mobile": safe_get_attr("father_mobile"),
            "Father Landline Code": safe_get_attr("father_landline_code"),
            "Father Landline": safe_get_attr("father_landline"),
            "Father Phone": safe_get_attr("father_phone"),
        })
        
        # Local Guardian's Details
        data.update({
            "Guardian Name": safe_get_attr("guardian_name"),
            "Guardian Relation": safe_get_attr("guardian_relation"),
            "Guardian Residential Address": safe_get_attr("guardian_residential_address"),
            "Guardian Organization": safe_get_attr("guardian_organization"),
            "Guardian Email": safe_get_attr("guardian_email"),
            "Guardian Mobile": safe_get_attr("guardian_mobile"),
            "Guardian Landline Code": safe_get_attr("guardian_landline_code"),
            "Guardian Landline": safe_get_attr("guardian_landline"),
            "Guardian Phone": safe_get_attr("guardian_phone"),
        })
        
        # Qualifying Examination
        data.update({
            "12th Year": safe_get_attr("twelfth_year"),
            "12th Board": safe_get_attr("twelfth_board"),
            "12th Roll Number": safe_get_attr("twelfth_roll_number"),
            "12th Institution": safe_get_attr("twelfth_institution"),
            "12th Percentage": safe_get_attr("twelfth_percentage"),
            "12th School": safe_get_attr("twelfth_school"),
            "Hindi Studied Upto": safe_get_attr("hindi_studied_upto"),
            "10th Board": safe_get_attr("tenth_board"),
            "10th Year": safe_get_attr("tenth_year"),
            "10th Percentage": safe_get_attr("tenth_percentage"),
            "10th School": safe_get_attr("tenth_school"),
            "Previous Qualification": safe_get_attr("previous_qualification"),
            "Graduation Details": safe_get_attr("graduation_details"),
        })
        
        # Other Information
        data.update({
            "DU Enrollment Number": safe_get_attr("du_enrollment_number"),
            "Hindi Medium Preference": safe_get_attr("hindi_medium_preference"),
        })
        
        # Category Certificate Details
        data.update({
            "Category Certificate Authority": safe_get_attr("category_certificate_authority"),
            "Category Certificate Number": safe_get_attr("category_certificate_number"),
            "Category Certificate Date": safe_get_attr("category_certificate_date"),
            "Disability Percentage": safe_get_attr("disability_percentage"),
            "Disability Type": safe_get_attr("disability_type"),
            "UDID Number": safe_get_attr("udid_number"),
        })
        
        # Legacy/Backward Compatibility
        data.update({
            "Course Applied": safe_get_attr("course_applied"),
            "Application Number": safe_get_attr("application_number"),
            "Enrollment Number": safe_get_attr("enrollment_number"),
            "Admission Date": safe_get_attr("date_of_admission"),
        })
    else:
        # Fill all fields with empty strings for consistency
        all_fields = [
            # Academic & Admission Details
            "Academic Session", "Course", "Admission Category", "Admission Category Other",
            "DU Portal Form Number", "CUET Score", "Total CUET Score", "College Roll No",
            "Date of Admission",
            # CUET Marks
            "CUET Subject 1", "CUET Total Score 1", "CUET Score Obtained 1",
            "CUET Subject 2", "CUET Total Score 2", "CUET Score Obtained 2",
            "CUET Subject 3", "CUET Total Score 3", "CUET Score Obtained 3",
            "CUET Subject 4", "CUET Total Score 4", "CUET Score Obtained 4",
            "CUET Subject 5", "CUET Total Score 5", "CUET Score Obtained 5",
            "CUET Subject 6", "CUET Total Score 6", "CUET Score Obtained 6",
            # Personal Details
            "First Name", "Middle Name", "Surname", "Date of Birth", "Gender", "Category",
            "Nationality", "Religion", "Blood Group", "Below Poverty Line", "Minority Category",
            "Annual Income",
            # Address
            "Permanent Address Line 1", "Permanent Address Line 2", "Permanent Address Line 3",
            "Permanent State", "Permanent Pincode", "Permanent Address",
            "Correspondence Address Line 1", "Correspondence Address Line 2", "Correspondence Address Line 3",
            "Correspondence State", "Correspondence Pincode", "Correspondence Address",
            "City", "State", "Pincode",
            # Contact
            "Email", "Phone Number", "Alternate Phone", "Emergency Contact Name", "Emergency Contact Phone",
            # Mother's Details
            "Mother Name", "Mother Occupation", "Mother Designation", "Mother Organization",
            "Mother Email", "Mother Mobile", "Mother Landline Code", "Mother Landline", "Mother Phone",
            # Father's Details
            "Father Name", "Father Occupation", "Father Designation", "Father Organization",
            "Father Email", "Father Mobile", "Father Landline Code", "Father Landline", "Father Phone",
            # Guardian's Details
            "Guardian Name", "Guardian Relation", "Guardian Residential Address", "Guardian Organization",
            "Guardian Email", "Guardian Mobile", "Guardian Landline Code", "Guardian Landline", "Guardian Phone",
            # Qualifying Examination
            "12th Year", "12th Board", "12th Roll Number", "12th Institution", "12th Percentage", "12th School",
            "Hindi Studied Upto", "10th Board", "10th Year", "10th Percentage", "10th School",
            "Previous Qualification", "Graduation Details",
            # Other Information
            "DU Enrollment Number", "Hindi Medium Preference",
            # Category Certificate
            "Category Certificate Authority", "Category Certificate Number", "Category Certificate Date",
            "Disability Percentage", "Disability Type", "UDID Number",
            # Legacy
            "Course Applied", "Application Number", "Enrollment Number", "Admission Date",
        ]
        for field in all_fields:
            data[field] = ""
    
    # Add counts
    data["Forms Count"] = db.query(AdmissionForm).filter(
        AdmissionForm.student_profile_id == profile.id
    ).count()
    data["Documents Count"] = db.query(StudentDocument).filter(
        StudentDocument.student_profile_id == profile.id
    ).count()
    
    return data

@router.get("/export/csv")
async def export_students_csv(
    student_name: Optional[str] = Query(None),
    roll_number: Optional[str] = Query(None),
    aadhar_number: Optional[str] = Query(None),
    phone_number: Optional[str] = Query(None),
    email: Optional[str] = Query(None),
    course_applied: Optional[str] = Query(None),
    academic_session: Optional[str] = Query(None),
    gender: Optional[str] = Query(None),
    category: Optional[str] = Query(None),
    city: Optional[str] = Query(None),
    state: Optional[str] = Query(None),
    is_verified: bool = Query(True, description="Only export verified students"),
    sort_by: str = Query("updated", description="Sort field: name, roll_number, aadhar, created, updated"),
    sort_order: str = Query("desc", description="Sort order: asc or desc"),
    db: Session = Depends(get_db),
    user: CurrentUser = Depends(RequireStaffOrAdmin),
):
    """Export student records to CSV with filtering and sorting - includes ALL form fields"""
    from fastapi.responses import StreamingResponse
    
    # Build query
    query = get_student_data_query(
        db, student_name, roll_number, aadhar_number,
        phone_number, email, course_applied, academic_session, gender, category, city, state,
        sort_by, sort_order, is_verified
    )
    
    profiles = query.all()
    
    # Create CSV in memory
    output = io.StringIO()
    writer = None
    
    if not profiles:
        # Return empty CSV with headers
        fieldnames = ["ID", "Student Name", "Roll Number", "Aadhar Number", "Created Date", "Updated Date"]
        writer = csv.DictWriter(output, fieldnames=fieldnames)
        writer.writeheader()
    else:
        for profile in profiles:
            data = get_student_export_data(profile, db)
            if writer is None:
                # Initialize writer with headers from first record
                fieldnames = list(data.keys())
                writer = csv.DictWriter(output, fieldnames=fieldnames)
                writer.writeheader()
            writer.writerow(data)
    
    output.seek(0)
    
    # Generate filename with timestamp
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    filename = f"students_export_{timestamp}.csv"
    
    return StreamingResponse(
        iter([output.getvalue()]),
        media_type="text/csv",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'}
    )

@router.get("/export/excel")
async def export_students_excel(
    student_name: Optional[str] = Query(None),
    roll_number: Optional[str] = Query(None),
    aadhar_number: Optional[str] = Query(None),
    phone_number: Optional[str] = Query(None),
    email: Optional[str] = Query(None),
    course_applied: Optional[str] = Query(None),
    academic_session: Optional[str] = Query(None),
    gender: Optional[str] = Query(None),
    category: Optional[str] = Query(None),
    city: Optional[str] = Query(None),
    state: Optional[str] = Query(None),
    is_verified: bool = Query(True, description="Only export verified students"),
    sort_by: str = Query("updated", description="Sort field: name, roll_number, aadhar, created, updated"),
    sort_order: str = Query("desc", description="Sort order: asc or desc"),
    db: Session = Depends(get_db),
    user: CurrentUser = Depends(RequireStaffOrAdmin),
):
    """Export student records to Excel with filtering and sorting - includes ALL form fields"""
    from fastapi.responses import StreamingResponse
    
    # Build query
    query = get_student_data_query(
        db, student_name, roll_number, aadhar_number,
        phone_number, email, course_applied, academic_session, gender, category, city, state,
        sort_by, sort_order, is_verified
    )
    
    profiles = query.all()
    
    if not profiles:
        # Return empty Excel file instead of 404 error
        output = io.BytesIO()
        df = pd.DataFrame(columns=["ID", "Student Name", "Roll Number", "Aadhar Number"])
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            df.to_excel(writer, sheet_name='Students', index=False)
        output.seek(0)
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"students_export_{timestamp}.xlsx"
        return StreamingResponse(
            output,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f'attachment; filename="{filename}"'}
        )
    
    # Build data for DataFrame
    all_data = []
    for profile in profiles:
        try:
            data = get_student_export_data(profile, db)
            # Remove ocr_provider and additional_info if present
            data.pop('ocr_provider', None)
            data.pop('OCR Provider', None)
            data.pop('additional_info', None)
            data.pop('Additional Info', None)
            all_data.append(data)
        except Exception as e:
            # Log error but continue with other profiles
            logger.error(f"Error getting export data for profile {profile.id}: {e}", exc_info=True)
            # Add basic profile data even if form data fails
            all_data.append({
                "ID": profile.id,
                "Student Name": profile.student_name or "",
                "Roll Number": profile.roll_number or "",
                "Aadhar Number": profile.aadhar_number or "",
                "Created Date": profile.created_date.strftime("%Y-%m-%d %H:%M:%S") if profile.created_date else "",
                "Updated Date": profile.updated_date.strftime("%Y-%m-%d %H:%M:%S") if profile.updated_date else "",
            })
    
    if not all_data:
        # Return empty Excel file with headers
        output = io.BytesIO()
        df = pd.DataFrame(columns=["ID", "Student Name", "Roll Number", "Aadhar Number", "Created Date", "Updated Date"])
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            df.to_excel(writer, sheet_name='Students', index=False)
        output.seek(0)
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"students_export_{timestamp}.xlsx"
        return StreamingResponse(
            output,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f'attachment; filename="{filename}"'}
        )
    
    # Create DataFrame
    try:
        df = pd.DataFrame(all_data)
        logger.info(f"Created DataFrame with {len(df)} rows and {len(df.columns)} columns")
    except Exception as e:
        logger.error(f"Error creating DataFrame: {e}", exc_info=True)
        # Fallback: create DataFrame with basic columns
        basic_data = []
        for profile in profiles:
            basic_data.append({
                "ID": profile.id,
                "Student Name": profile.student_name or "",
                "Roll Number": profile.roll_number or "",
                "Aadhar Number": profile.aadhar_number or "",
            })
        df = pd.DataFrame(basic_data)
        logger.info(f"Created fallback DataFrame with {len(df)} rows")
    
    # Ensure DataFrame is not empty
    if df.empty:
        logger.warning("DataFrame is empty after creation, recreating from first data dict")
        # If DataFrame is empty, create one with headers from first data dict
        if all_data:
            df = pd.DataFrame([all_data[0]])
            logger.info(f"Recreated DataFrame with {len(df)} rows from first data dict")
        else:
            df = pd.DataFrame(columns=["ID", "Student Name", "Roll Number", "Aadhar Number"])
    
    # Remove columns if they still exist
    columns_to_drop = ['ocr_provider', 'OCR Provider', 'additional_info', 'Additional Info']
    df = df.drop(columns=[col for col in columns_to_drop if col in df.columns], errors='ignore')
    
    logger.info(f"Final DataFrame shape: {df.shape} (rows x columns)")
    
    # Create Excel in memory
    output = io.BytesIO()
    try:
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
        
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            df.to_excel(writer, sheet_name='Students', index=False)
            
            # Get the worksheet to format it
            worksheet = writer.sheets['Students']
            
            # Define styles
            header_font = Font(name='Calibri', size=11, bold=True, color='FFFFFF')
            header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
            header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            cell_font = Font(name='Calibri', size=10)
            cell_alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
            border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            
            # Format header row (row 1)
            for idx, col in enumerate(df.columns, start=1):
                cell = worksheet.cell(row=1, column=idx)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment
                cell.border = border
            
            # Format data rows
            for row_idx in range(2, len(df) + 2):  # Start from row 2 (after header)
                for col_idx in range(1, len(df.columns) + 1):
                    cell = worksheet.cell(row=row_idx, column=col_idx)
                    cell.font = cell_font
                    cell.alignment = cell_alignment
                    cell.border = border
            
            # Auto-adjust column widths with better calculation
            for idx, col in enumerate(df.columns, start=1):
                try:
                    # Calculate max length in column data
                    if len(df) > 0:
                        max_data_length = df[col].astype(str).map(len).max()
                    else:
                        max_data_length = 0
                    
                    # Get header length
                    header_length = len(str(col))
                    
                    # Calculate width: max of data and header, with padding
                    max_length = max(max_data_length, header_length)
                    
                    # Set width with better logic:
                    # - Minimum width: 10
                    # - Maximum width: 60 (increased from 50)
                    # - Add padding: 3 characters
                    # - For very short columns, use a reasonable minimum
                    if max_length < 10:
                        adjusted_width = 12
                    elif max_length > 57:
                        adjusted_width = 60
                    else:
                        adjusted_width = max_length + 3
                    
                    worksheet.column_dimensions[get_column_letter(idx)].width = adjusted_width
                except Exception as e:
                    # If column formatting fails, continue with other columns
                    logger.warning(f"Error formatting column {col}: {e}")
            
            # Set row heights
            worksheet.row_dimensions[1].height = 25  # Header row height
            for row_idx in range(2, len(df) + 2):
                worksheet.row_dimensions[row_idx].height = 18  # Data row height
            
            # Freeze header row
            worksheet.freeze_panes = 'A2'
            
            # Enable filter on header row
            worksheet.auto_filter.ref = worksheet.dimensions
        
        output.seek(0)
        file_size = len(output.getvalue())
        logger.info(f"Excel file created successfully, size: {file_size} bytes, rows: {len(df)}, columns: {len(df.columns)}")
        
        if file_size < 1000:  # Very small file might be empty
            logger.warning(f"Excel file is very small ({file_size} bytes), might be empty")
    except Exception as e:
        logger.error(f"Error creating Excel file: {e}", exc_info=True)
        raise HTTPException(status_code=500, detail=f"Error creating Excel file: {str(e)}")
    
    output.seek(0)
    
    # Generate filename with timestamp
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    filename = f"students_export_{timestamp}.xlsx"
    
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'}
    )


@router.get("/export/json")
async def export_students_json(
    student_name: Optional[str] = Query(None),
    roll_number: Optional[str] = Query(None),
    aadhar_number: Optional[str] = Query(None),
    phone_number: Optional[str] = Query(None),
    email: Optional[str] = Query(None),
    course_applied: Optional[str] = Query(None),
    academic_session: Optional[str] = Query(None),
    gender: Optional[str] = Query(None),
    category: Optional[str] = Query(None),
    city: Optional[str] = Query(None),
    state: Optional[str] = Query(None),
    is_verified: bool = Query(True, description="Only export verified students"),
    sort_by: str = Query("updated", description="Sort field: name, roll_number, aadhar, created, updated"),
    sort_order: str = Query("desc", description="Sort order: asc or desc"),
    db: Session = Depends(get_db),
    user: CurrentUser = Depends(RequireStaffOrAdmin),
):
    """Export student records to JSON with filtering and sorting - includes ALL form fields"""
    from fastapi.responses import StreamingResponse
    import json
    
    # Build query
    query = get_student_data_query(
        db, student_name, roll_number, aadhar_number,
        phone_number, email, course_applied, academic_session, gender, category, city, state,
        sort_by, sort_order, is_verified
    )
    
    profiles = query.all()
    
    if not profiles:
        # Return empty JSON array instead of 404 error
        json_output = json.dumps([], indent=2, ensure_ascii=False)
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"students_export_{timestamp}.json"
        return StreamingResponse(
            iter([json_output]),
            media_type="application/json",
            headers={"Content-Disposition": f'attachment; filename="{filename}"'}
        )
    
    # Build data list
    all_data = []
    for profile in profiles:
        data = get_student_export_data(profile, db)
        all_data.append(data)
    
    # Create JSON string
    json_output = json.dumps(all_data, indent=2, ensure_ascii=False, default=str)
    
    # Generate filename with timestamp
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    filename = f"students_export_{timestamp}.json"
    
    return StreamingResponse(
        iter([json_output]),
        media_type="application/json",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'}
    )


@router.get("/export/pdf")
async def export_students_pdf(
    student_name: Optional[str] = Query(None),
    roll_number: Optional[str] = Query(None),
    aadhar_number: Optional[str] = Query(None),
    phone_number: Optional[str] = Query(None),
    email: Optional[str] = Query(None),
    course_applied: Optional[str] = Query(None),
    academic_session: Optional[str] = Query(None),
    gender: Optional[str] = Query(None),
    category: Optional[str] = Query(None),
    city: Optional[str] = Query(None),
    state: Optional[str] = Query(None),
    is_verified: bool = Query(True, description="Only export verified students"),
    sort_by: str = Query("updated", description="Sort field: name, roll_number, aadhar, created, updated"),
    sort_order: str = Query("desc", description="Sort order: asc or desc"),
    db: Session = Depends(get_db),
    user: CurrentUser = Depends(RequireStaffOrAdmin),
):
    """Export student records to PDF with filtering and sorting"""
    from fastapi.responses import StreamingResponse
    
    # Build query
    query = get_student_data_query(
        db, student_name, roll_number, aadhar_number,
        phone_number, email, course_applied, academic_session, gender, category, city, state,
        sort_by, sort_order, is_verified
    )
    
    profiles = query.all()
    
    # Build data list
    all_data = []
    for profile in profiles:
        data = get_student_export_data(profile, db)
        all_data.append(data)
    
    # Handle empty results
    if not all_data:
        # Return empty PDF
        try:
            from reportlab.lib.pagesizes import A4
            from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer
            from reportlab.lib.styles import getSampleStyleSheet
            
            output = io.BytesIO()
            doc = SimpleDocTemplate(output, pagesize=A4)
            styles = getSampleStyleSheet()
            elements = [
                Paragraph("SRCC Student Records Export", styles['Heading1']),
                Spacer(1, 20),
                Paragraph("No students found matching the criteria.", styles['Normal'])
            ]
            doc.build(elements)
            output.seek(0)
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            filename = f"students_export_{timestamp}.pdf"
            return StreamingResponse(
                output,
                media_type="application/pdf",
                headers={"Content-Disposition": f'attachment; filename="{filename}"'}
            )
        except ImportError:
            raise HTTPException(
                status_code=500,
                detail="PDF export requires reportlab library. Install with: pip install reportlab"
            )
    
    # Try to use reportlab for PDF generation
    try:
        from reportlab.lib import colors
        from reportlab.lib.pagesizes import A4, landscape
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.lib.units import inch, cm
        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, PageBreak
        from reportlab.lib.enums import TA_CENTER, TA_LEFT
        
        output = io.BytesIO()
        doc = SimpleDocTemplate(output, pagesize=landscape(A4), topMargin=0.5*inch, bottomMargin=0.5*inch)
        
        styles = getSampleStyleSheet()
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=18,
            alignment=TA_CENTER,
            spaceAfter=20
        )
        
        elements = []
        
        # Title
        elements.append(Paragraph("SRCC Student Records Export", title_style))
        elements.append(Paragraph(f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}", styles['Normal']))
        elements.append(Paragraph(f"Total Records: {len(all_data)}", styles['Normal']))
        elements.append(Spacer(1, 20))
        
        # Key fields for summary table
        summary_headers = ['#', 'Student Name', 'Roll Number', 'Course', 'Category', 'Phone', 'Email']
        table_data = [summary_headers]
        
        for idx, student in enumerate(all_data, 1):
            row = [
                str(idx),
                str(student.get('Student Name', '-'))[:30],
                str(student.get('Roll Number', '-'))[:15],
                str(student.get('Course', '-'))[:20],
                str(student.get('Category', '-'))[:10],
                str(student.get('Phone Number', '-'))[:15],
                str(student.get('Email', '-'))[:25],
            ]
            table_data.append(row)
        
        # Create table
        col_widths = [0.4*inch, 2*inch, 1.2*inch, 1.5*inch, 0.8*inch, 1.2*inch, 2*inch]
        table = Table(table_data, colWidths=col_widths, repeatRows=1)
        
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.Color(0.1, 0.23, 0.43)),  # Navy blue header
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 10),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('TOPPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.white),
            ('TEXTCOLOR', (0, 1), (-1, -1), colors.black),
            ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
            ('FONTSIZE', (0, 1), (-1, -1), 8),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.Color(0.95, 0.95, 0.95)]),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('LEFTPADDING', (0, 0), (-1, -1), 6),
            ('RIGHTPADDING', (0, 0), (-1, -1), 6),
        ]))
        
        elements.append(table)
        
        doc.build(elements)
        output.seek(0)
        
        # Generate filename with timestamp
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"students_export_{timestamp}.pdf"
        
        return StreamingResponse(
            output,
            media_type="application/pdf",
            headers={"Content-Disposition": f'attachment; filename="{filename}"'}
        )
        
    except ImportError:
        # Fallback: return error if reportlab not installed
        raise HTTPException(
            status_code=500, 
            detail="PDF export requires reportlab library. Install with: pip install reportlab"
        )


# Field mapping for import - maps CSV/Excel column names to database fields
IMPORT_FIELD_MAPPING = {
    # Profile fields
    'student name': 'student_name',
    'roll number': 'roll_number',
    'aadhar number': 'aadhar_number',
    # Academic & Admission Details
    'academic session': 'academic_session',
    'course': 'course',
    'admission category': 'admission_category',
    'admission category other': 'admission_category_other',
    'du portal form number': 'du_portal_form_number',
    'cuet score': 'cuet_score',
    'total cuet score': 'cuet_total_score',
    'college roll no': 'college_roll_no',
    'date of admission': 'date_of_admission',
    # CUET Marks
    'cuet subject 1': 'cuet_subject_1', 'cuet total score 1': 'cuet_total_score_1', 'cuet score obtained 1': 'cuet_score_obtained_1',
    'cuet subject 2': 'cuet_subject_2', 'cuet total score 2': 'cuet_total_score_2', 'cuet score obtained 2': 'cuet_score_obtained_2',
    'cuet subject 3': 'cuet_subject_3', 'cuet total score 3': 'cuet_total_score_3', 'cuet score obtained 3': 'cuet_score_obtained_3',
    'cuet subject 4': 'cuet_subject_4', 'cuet total score 4': 'cuet_total_score_4', 'cuet score obtained 4': 'cuet_score_obtained_4',
    'cuet subject 5': 'cuet_subject_5', 'cuet total score 5': 'cuet_total_score_5', 'cuet score obtained 5': 'cuet_score_obtained_5',
    'cuet subject 6': 'cuet_subject_6', 'cuet total score 6': 'cuet_total_score_6', 'cuet score obtained 6': 'cuet_score_obtained_6',
    # Personal Details
    'first name': 'first_name', 'middle name': 'middle_name', 'surname': 'surname',
    'date of birth': 'date_of_birth', 'gender': 'gender', 'category': 'category',
    'nationality': 'nationality', 'religion': 'religion', 'blood group': 'blood_group',
    'below poverty line': 'below_poverty_line', 'minority category': 'minority_category', 'annual income': 'annual_income',
    # Address
    'permanent address line 1': 'permanent_address_line1', 'permanent address line 2': 'permanent_address_line2',
    'permanent address line 3': 'permanent_address_line3', 'permanent state': 'permanent_state',
    'permanent pincode': 'permanent_pincode', 'permanent address': 'permanent_address',
    'correspondence address line 1': 'correspondence_address_line1', 'correspondence address line 2': 'correspondence_address_line2',
    'correspondence address line 3': 'correspondence_address_line3', 'correspondence state': 'correspondence_state',
    'correspondence pincode': 'correspondence_pincode', 'correspondence address': 'correspondence_address',
    'city': 'city', 'state': 'state', 'pincode': 'pincode',
    # Contact
    'email': 'email', 'phone number': 'phone_number', 'alternate phone': 'alternate_phone',
    'emergency contact name': 'emergency_contact_name', 'emergency contact phone': 'emergency_contact_phone',
    # Mother's Details
    'mother name': 'mother_name', 'mother occupation': 'mother_occupation', 'mother designation': 'mother_designation',
    'mother organization': 'mother_organization', 'mother email': 'mother_email', 'mother mobile': 'mother_mobile',
    'mother landline code': 'mother_landline_code', 'mother landline': 'mother_landline', 'mother phone': 'mother_phone',
    # Father's Details
    'father name': 'father_name', 'father occupation': 'father_occupation', 'father designation': 'father_designation',
    'father organization': 'father_organization', 'father email': 'father_email', 'father mobile': 'father_mobile',
    'father landline code': 'father_landline_code', 'father landline': 'father_landline', 'father phone': 'father_phone',
    # Guardian's Details
    'guardian name': 'guardian_name', 'guardian relation': 'guardian_relation',
    'guardian residential address': 'guardian_residential_address', 'guardian organization': 'guardian_organization',
    'guardian email': 'guardian_email', 'guardian mobile': 'guardian_mobile',
    'guardian landline code': 'guardian_landline_code', 'guardian landline': 'guardian_landline', 'guardian phone': 'guardian_phone',
    # Qualifying Examination
    '12th year': 'twelfth_year', '12th board': 'twelfth_board', '12th roll number': 'twelfth_roll_number',
    '12th institution': 'twelfth_institution', '12th percentage': 'twelfth_percentage', '12th school': 'twelfth_school',
    'hindi studied upto': 'hindi_studied_upto', '10th board': 'tenth_board', '10th year': 'tenth_year',
    '10th percentage': 'tenth_percentage', '10th school': 'tenth_school',
    'previous qualification': 'previous_qualification', 'graduation details': 'graduation_details',
    # Other Information
    'du enrollment number': 'du_enrollment_number', 'hindi medium preference': 'hindi_medium_preference',
    # Category Certificate
    'category certificate authority': 'category_certificate_authority', 'category certificate number': 'category_certificate_number',
    'category certificate date': 'category_certificate_date', 'disability percentage': 'disability_percentage',
    'disability type': 'disability_type', 'udid number': 'udid_number',
    # Legacy
    'course applied': 'course_applied', 'application number': 'application_number',
    'enrollment number': 'enrollment_number', 'admission date': 'date_of_admission',
}

def normalize_field_name(field_name: str) -> str:
    """Normalize field name to lowercase and strip whitespace"""
    return field_name.lower().strip()

@router.post("/import/csv")
async def import_students_csv(
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
    user: CurrentUser = Depends(RequireStaffOrAdmin),
):
    """Import student records from CSV file - requires all form fields"""
    if not file.filename.endswith('.csv'):
        raise HTTPException(status_code=400, detail="File must be a CSV file")
    
    # Read CSV
    contents = await file.read()
    csv_data = io.StringIO(contents.decode('utf-8'))
    reader = csv.DictReader(csv_data)
    
    # Get column names and normalize them
    original_columns = reader.fieldnames or []
    normalized_columns = {normalize_field_name(col): col for col in original_columns}
    
    imported = 0
    errors = []
    
    for idx, row in enumerate(reader, start=2):  # Start at 2 (1 is header)
        try:
            # Required field: Student Name
            student_name_key = None
            for key in ['student name', 'student_name', 'name']:
                if key in normalized_columns:
                    student_name_key = normalized_columns[key]
                    break
            
            if not student_name_key or not row.get(student_name_key):
                errors.append(f"Row {idx}: Student Name is required")
                continue
            
            student_name = str(row[student_name_key]).strip()
            
            # Get aadhar_number for matching
            aadhar_key = None
            for key in ['aadhar number', 'aadhar_number', 'aadhar']:
                if key in normalized_columns:
                    aadhar_key = normalized_columns[key]
                    break
            aadhar_number = str(row.get(aadhar_key, '')).strip() if aadhar_key and row.get(aadhar_key) else None
            
            # Get or create student profile
            from backend.api.routes.students import get_or_create_student_profile
            profile = get_or_create_student_profile(db, student_name, aadhar_number=aadhar_number)
            
            # Map all CSV columns to form fields
            form_data = {}
            for norm_col, orig_col in normalized_columns.items():
                value = row.get(orig_col, '')
                if value and str(value).strip():
                    # Map normalized column name to database field
                    db_field = IMPORT_FIELD_MAPPING.get(norm_col)
                    if db_field and hasattr(AdmissionForm, db_field):
                        form_data[db_field] = str(value).strip()
            
            # Always create a form record with all available data
            form = AdmissionForm(
                filename=f"imported_{profile.id}_{idx}.csv",
                file_path="imported",
                ocr_provider="csv_import",
                status="verified",
                student_profile_id=profile.id,
                student_name=student_name,
                extracted_data={"imported_from_csv": True, "row_number": idx},
                **{k: v for k, v in form_data.items() if hasattr(AdmissionForm, k)}
            )
            db.add(form)
            db.commit()
            imported += 1
                    
        except Exception as e:
            errors.append(f"Row {idx}: {str(e)}")
            db.rollback()
    
    return {
        "message": f"Import completed",
        "imported": imported,
        "errors": errors,
        "error_count": len(errors)
    }

@router.post("/import/excel")
async def import_students_excel(
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
    user: CurrentUser = Depends(RequireStaffOrAdmin),
):
    """Import student records from Excel file - requires all form fields"""
    if not (file.filename.endswith('.xlsx') or file.filename.endswith('.xls')):
        raise HTTPException(status_code=400, detail="File must be an Excel file (.xlsx or .xls)")
    
    # Read Excel
    contents = await file.read()
    excel_data = io.BytesIO(contents)
    
    try:
        df = pd.read_excel(excel_data, sheet_name=0)
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Error reading Excel file: {str(e)}")
    
    # Normalize column names
    normalized_columns = {normalize_field_name(col): col for col in df.columns}
    
    imported = 0
    errors = []
    
    for idx, row in df.iterrows():
        try:
            # Required field: Student Name
            student_name_key = None
            for key in ['student name', 'student_name', 'name']:
                if key in normalized_columns:
                    student_name_key = normalized_columns[key]
                    break
            
            if not student_name_key or pd.isna(row.get(student_name_key)) or not str(row.get(student_name_key)).strip():
                errors.append(f"Row {idx + 2}: Student Name is required")  # +2 because Excel is 1-indexed and has header
                continue
            
            student_name = str(row[student_name_key]).strip()
            
            # Get aadhar_number for matching
            aadhar_key = None
            for key in ['aadhar number', 'aadhar_number', 'aadhar']:
                if key in normalized_columns:
                    aadhar_key = normalized_columns[key]
                    break
            aadhar_value = row.get(aadhar_key) if aadhar_key else None
            aadhar_number = str(aadhar_value).strip() if not pd.isna(aadhar_value) and aadhar_value else None
            
            # Get or create student profile
            from backend.api.routes.students import get_or_create_student_profile
            profile = get_or_create_student_profile(db, student_name, aadhar_number=aadhar_number)
            
            # Map all Excel columns to form fields
            form_data = {}
            for norm_col, orig_col in normalized_columns.items():
                value = row.get(orig_col)
                if not pd.isna(value) and value and str(value).strip():
                    # Map normalized column name to database field
                    db_field = IMPORT_FIELD_MAPPING.get(norm_col)
                    if db_field and hasattr(AdmissionForm, db_field):
                        form_data[db_field] = str(value).strip()
            
            # Always create a form record with all available data
            form = AdmissionForm(
                filename=f"imported_{profile.id}_{idx + 2}.xlsx",
                file_path="imported",
                ocr_provider="excel_import",
                status="verified",
                student_profile_id=profile.id,
                student_name=student_name,
                extracted_data={"imported_from_excel": True, "row_number": idx + 2},
                **{k: v for k, v in form_data.items() if hasattr(AdmissionForm, k)}
            )
            db.add(form)
            db.commit()
            imported += 1
                    
        except Exception as e:
            errors.append(f"Row {idx + 2}: {str(e)}")
            db.rollback()
    
    return {
        "message": f"Import completed",
        "imported": imported,
        "errors": errors,
        "error_count": len(errors)
    }
